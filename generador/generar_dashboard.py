# -*- coding: utf-8 -*-
"""Generador del Dashboard Forecast (Electroestrada).
Uso: python generar_dashboard.py [ruta_xlsm] [salida_html]"""
import sys, os, json, glob, datetime, re
from datetime import timedelta
import openpyxl
from openpyxl.utils import column_index_from_string as ci

def _total(s):
    m=re.search(r'Total:\s*([\d\.]+)', s or "")
    return int(m.group(1).replace(".","")) if m else None
def _usd(s):
    s=s or ""
    def num(tag):
        m=re.search(tag+r'\s*:?\s*U\$D\s*([\d\.]+,\d{2}|[\d\.]+)', s)
        if not m: return None
        v=m.group(1).replace(".","").replace(",",".")
        try: return float(v)
        except: return None
    pi=num("PI"); cii=num("CI")
    if pi is None and cii is None:
        m=re.match(r'\s*([\d\.]+)\s*$', s)
        if m:
            try: pi=float(m.group(1))
            except: pass
    return [pi,cii]

def find_latest(folder):
    cands=[p for p in glob.glob(os.path.join(folder,"*.xlsm")) if not os.path.basename(p).startswith("~$")]
    if not cands: raise SystemExit("No se encontro ningun .xlsm en "+folder)
    return max(cands, key=os.path.getmtime)

def extract(path):
    wb=openpyxl.load_workbook(path, read_only=True, data_only=True, keep_links=False)
    ws=wb["Forecast"]
    BS=88; STRIDE=14
    all_rows=[list(x) for x in ws.iter_rows(min_row=1,max_row=6000,max_col=270,values_only=True)]
    _rl=all_rows[2] if len(all_rows)>2 else []
    real_labels=[(_rl[cc] if cc<len(_rl) else None) for cc in range(ci("AB")-1, ci("AB")-1+12)]
    r1=all_rows[0]
    months=[str(r1[BS+STRIDE*k]) for k in range(12)]
    def I(v):
        try: return int(round(float(v)))
        except: return None
    def F1(v):
        try: return round(float(v),1)
        except: return None
    def N(v):
        try: f=float(v)
        except: return None
        r=round(f,3)
        return int(r) if r==int(r) else r
    UN=[];GA=[];CAT=[]
    def idx(lst,v):
        v="" if v is None else str(v)
        if v not in lst: lst.append(v)
        return lst.index(v)
    rows=[]
    realmap={}
    for row in all_rows[3:]:
        cod=row[2]
        if not (cod and str(cod).strip()): continue
        ui=idx(UN,row[0]); gi=idx(GA,row[1]); cti=idx(CAT,row[5])
        sa=N(row[ci("BN")-1]); ma=F1(row[ci("BO")-1])
        flat=[]; prev=ma
        for k in range(12):
            b=BS+STRIDE*k; bn=BS+STRIDE*(k+1)
            si=N(row[b+5-1]); mf=F1(row[b+7-1]); comp=N(row[b+10-1]); ven=N(row[b+1-1])
            ni=bn+5-1; sf=N(row[ni]) if ni<len(row) else None
            if sf is None and si is not None and comp is not None and ven is not None: sf=si+comp-ven
            flat += [si, prev, comp, ven, sf, mf]; prev=mf
        ideal=F1(row[ci("H")-1])
        vpx=[]
        for k in range(12):
            b=BS+STRIDE*k
            vpx += [N(row[b+0-1]), N(row[b+4-1])]  # venta base, venta proy USD/$
        try: an=round(float(row[ci("AN")-1]),4)
        except: an=None
        try: pp=round(float(row[ci("AS")-1]),4)
        except: pp=None
        _rr=[]
        for cc in range(ci("AB")-1, ci("AB")-1+12):
            try: _rr.append(abs(float(row[cc])))
            except: _rr.append(None)
        realmap[str(cod).strip()]=_rr
        rows.append([ui,gi,str(cod).strip(),cti,sa,ma]+flat+[ideal]+vpx+[an]+[pp])
    # IMPO por GA
    def estado_of(H1,b):
        for off in (8,9):
            v=H1[b-1+off]
            if v:
                u=str(v).upper()
                if "CONFIRM" in u: return "confirmado"
                if "PROYECT" in u: return "proyectado"
        return "-"
    heads=[]
    for i in range(3,len(all_rows)):
        row=all_rows[i]; b=row[1]; c=row[2]
        if (b and str(b).strip()) and not (c and str(c).strip()):
            heads.append((i,str(b).strip()))
    groups={}
    j=0
    while j<len(heads):
        gidx,ga=heads[j]; grp=[gidx]; k=j+1
        while k<len(heads) and heads[k][0]==heads[k-1][0]+1 and heads[k][1]==ga:
            grp.append(heads[k][0]); k+=1
        if ga not in groups: groups[ga]=grp
        j=k
    impo={}
    for ga,grp in groups.items():
        H1=all_rows[grp[0]]; H2=all_rows[grp[1]] if len(grp)>1 else H1
        arr=[]
        for kk in range(12):
            b=BS+STRIDE*kk
            num=H1[b-1+10]; fe=H1[b-1+12]
            disp=""; eta=""
            if hasattr(fe,"strftime"):
                disp=fe.strftime("%d/%m/%Y"); eta=(fe-timedelta(days=10)).strftime("%d/%m/%Y")
            elif fe: disp=str(fe)
            uc=_usd(str(H2[b-1+12] or "")); up=_usd(str(H2[b-1+9] or ""))
            arr.append({"mes":months[kk],"estado":estado_of(H1,b),
                "impo":(str(num).strip() if num else ""),
                "disp":disp,"eta":eta,
                "cant":_total(str(H2[b-1+10] or "")),"pi":uc[0],"ci":uc[1],
                "cant_proy":_total(str(H2[b-1+8] or "")),"pi_proy":up[0],"ci_proy":up[1]})
        impo[ga]=arr
    # ---- Estacionalidad / Evento especial / TC por mes (row2 de cada bloque) ----
    season=[]
    for k in range(12):
        b=BS+STRIDE*k
        def r2(off): return all_rows[1][b+off-1]
        season.append([F1(r2(9)), F1(r2(12)), F1(r2(4))])  # estac, evento, TC
    # ---- Hoja VP-$ U$D (proyeccion comercial) ----
    vp=None
    try:
        vs=wb["VP-$ U$D"]
        VR=[list(x) for x in vs.iter_rows(min_row=1,max_row=44,max_col=55,values_only=True)]
        def vg(r,c): return VR[r-1][c-1] if (r-1<len(VR) and c-1<len(VR[r-1])) else None
        def numf(x):
            try: return round(float(x),2)
            except: return None
        def numi(x):
            try: return int(round(float(x)))
            except: return None
        vmonths=[str(vg(1,2+4*m)) for m in range(12)]
        vratio=[(str(vg(1,2+4*m+2)).strip() if vg(1,2+4*m+2) else "") for m in range(12)]
        vtotal=[numf(vg(3,2+4*m)) for m in range(12)]
        vtc=[]
        for m in range(12):
            val=None
            for off in (1,2,3):
                v=numf(vg(2,2+4*m+off))
                if v is not None: val=v; break
            vtc.append(val)
        vrows=[]
        for r in range(6,44):
            ga=vg(r,1)
            if not (ga and str(ga).strip()): continue
            ser=[]
            for m in range(12):
                b=2+4*m
                ser.append([numi(vg(r,b)), numi(vg(r,b+1)), numi(vg(r,b+2))])
            trend=[numf(vg(r,51)), numf(vg(r,52))]
            vrows.append([str(ga).strip(), ser, trend])
        vp={"months":vmonths,"ratio":vratio,"total":vtotal,"tc":vtc,"rows":vrows}
    except Exception as e:
        vp={"error":str(e)}
    # fecha/hora del stock (BN2 + BO2)
    bn=all_rows[1][ci("BN")-1]; bo=all_rows[1][ci("BO")-1]
    dias=["Lunes","Martes","Miercoles","Jueves","Viernes","Sabado","Domingo"]
    mes_es=["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
    if hasattr(bn,"strftime"):
        ts="%s %02d de %s %d"%(dias[bn.weekday()],bn.day,mes_es[bn.month-1],bn.year)
    else: ts=str(bn or "")
    hora=""
    if bo:
        mm=re.search(r"(\d{1,2}:\d{2})",str(bo)); hora=mm.group(1) if mm else ""
    stock_ts=ts+(" / "+hora+" hs" if hora else "")
    return {"months":months,"UN":UN,"GA":GA,"CAT":CAT,"rows":rows,"impo":impo,"stock_ts":stock_ts,"vp":vp,"season":season,
            "real":realmap,"real_labels":real_labels,
            "src":os.path.basename(path),"gen":datetime.datetime.now().strftime("%d/%m/%Y %H:%M")}

def load_costos(path):
    try:
        wb=openpyxl.load_workbook(path, read_only=True, data_only=True, keep_links=False)
        gs=wb["General"]; d={}
        def nf(x):
            try: return round(float(x),4)
            except: return None
        for r in gs.iter_rows(min_row=4, max_row=6000, max_col=30, values_only=True):
            cod=r[1]
            if cod and str(cod).strip():
                cur=r[ci("I")-1]; cur="USD" if (cur and "USD" in str(cur).upper()) else "$"
                d[str(cod).strip()]=[nf(r[ci("AC")-1]), nf(r[ci("J")-1]), nf(r[ci("F")-1]), cur, nf(r[ci("Q")-1])]
        return d
    except Exception:
        return {}

def find_costos(folder):
    import glob as _g
    for p in _g.glob(os.path.join(folder,"*ostos*.xlsm")):
        if not os.path.basename(p).startswith("~$"): return p
    return None

_MES={'ene':1,'feb':2,'mar':3,'abr':4,'may':5,'jun':6,'jul':7,'ago':8,'sep':9,'sept':9,'oct':10,'nov':11,'dic':12}
def _parse_label(s):
    if isinstance(s,(datetime.datetime, datetime.date)): return (s.year, s.month)
    s=str(s or "").strip().lower().replace('.','')
    m=re.match(r'([a-zñ]+)\s*[-_/ ]\s*(\d{2,4})', s)
    if not m: return None
    tok=m.group(1); mo=_MES.get(tok[:4]) or _MES.get(tok[:3])
    if not mo: return None
    y=int(m.group(2)); y=2000+y if y<100 else y
    return (y,mo)
def _parse_fname(fn):
    m=re.search(r'forecast\s+(\d{1,2})\s*[-_ ]\s*(\d{2,4})', fn.lower())
    if not m: return None
    mo=int(m.group(1)); y=int(m.group(2)); y=2000+y if y<100 else y
    return (y,mo)
_MESF={'enero':1,'febrero':2,'marzo':3,'abril':4,'mayo':5,'junio':6,'julio':7,'agosto':8,'septiembre':9,'setiembre':9,'octubre':10,'noviembre':11,'diciembre':12}
def _month_of(val):
    if isinstance(val,(datetime.datetime,datetime.date)): return val.month
    s=str(val or "").strip().lower().replace('.','')
    if not s: return None
    if s in _MESF: return _MESF[s]
    for name,mo in _MESF.items():
        if s.startswith(name[:4]): return mo
    pr=_parse_label(s)
    return pr[1] if pr else None
def extract_vp_archive(path, month):
    wb=openpyxl.load_workbook(path, read_only=True, data_only=True, keep_links=False)
    ws=wb["Forecast"]
    rr=[list(x) for x in ws.iter_rows(min_row=1, max_row=6000, max_col=200, values_only=True)]
    r1=rr[0] if rr else []; r2=rr[1] if len(rr)>1 else []; r3=rr[2] if len(rr)>2 else []
    targets=[]
    for hdr in (r1,r2):
        for cix,val in enumerate(hdr):
            if _month_of(val)==month: targets.append(cix)
    targets=sorted(set(targets))
    vpcol=None
    for t in targets:
        for cix in range(t, min(t+16,len(r3))):
            h=str(r3[cix] or "").lower()
            if ("venta proy" in h) and ("unidad" in h): vpcol=cix; break
        if vpcol is not None: break
    if vpcol is None: return {}
    mixcol=None
    for hdr in (r1,r2,r3):
        for cix,val in enumerate(hdr):
            if str(val or "").strip().upper()=="MIX": mixcol=cix; break
        if mixcol is not None: break
    d={}
    for row in rr[5:]:
        cod=row[2] if len(row)>2 else None
        if not (cod and str(cod).strip()): continue
        try:
            v=float(row[vpcol]); v=round(v,3); v=int(v) if v==int(v) else v
        except: v=None
        mx=None
        if mixcol is not None:
            try: mx=round(float(row[mixcol]),4)
            except: mx=None
        d[str(cod).strip()]=[v,mx]
    return d
def read_realusd(folder):
    import glob as _g
    cand=None
    for pat in ["Venta_real*.xls*","Venta real*.xls*","V.R*.xls*","VR *.xls*","V R *.xls*"]:
        for pp in _g.glob(os.path.join(folder,pat)):
            if not os.path.basename(pp).startswith("~$"): cand=pp; break
        if cand: break
    if not cand:
        print("  Venta real: no se encontro archivo 'Venta_real*.xlsx' en la carpeta"); return {}
    try:
        wb=openpyxl.load_workbook(cand, read_only=True, data_only=True, keep_links=False)
        ws=wb.active
        rr=[list(x) for x in ws.iter_rows(min_row=1,max_row=8000,max_col=40,values_only=True)]
        r1=rr[0] if rr else []; r2=rr[1] if len(rr)>1 else []
        mcol={}; cur=None
        for c in range(len(r1)):
            mo=_month_of(r1[c]) if r1[c] else None
            if mo: cur=mo
            sub=str((r2[c] if c<len(r2) else "") or "").strip().lower()
            if ("usd" in sub) and cur: mcol[cur]=c  # acepta "USD" y "$-USD"
        codcol=1
        for c in range(len(r1)):
            if "digo" in str(r1[c] or "").lower(): codcol=c; break
        d={}
        for row in rr[2:]:
            cod=row[codcol] if codcol<len(row) else None
            if not (cod and str(cod).strip()) or str(cod).strip().upper()=="TOTAL": continue
            m={}
            for mo,c in mcol.items():
                try: m[mo]=float(row[c])
                except: pass
            if m: d[str(cod).strip()]=m
        print("  Venta real: %s -> %d codigos, meses %s"%(os.path.basename(cand),len(d),sorted(mcol.keys())))
        return d
    except Exception as e:
        print("  Venta real: error",e); return {}

def build_historico(folder, real_map, real_labels, cache_path, realusd):
    cache={}
    if os.path.exists(cache_path):
        try: cache=json.load(open(cache_path,encoding="utf-8"))
        except: cache={}
    for pth in glob.glob(os.path.join(folder,"Forecast *")):
        bn=os.path.basename(pth)
        if bn.startswith("~$"): continue
        if not bn.lower().endswith((".xlsx",".xlsm")): continue
        pr=_parse_fname(bn)
        if not pr: continue
        key="%04d-%02d"%pr
        if key in cache: continue
        print("  Historico: leyendo",bn,"(una sola vez, puede tardar)...")
        try: cache[key]=extract_vp_archive(pth, pr[1])
        except Exception as e: print("    error:",e); continue
    try: json.dump(cache, open(cache_path,"w",encoding="utf-8"), ensure_ascii=False, separators=(",",":"))
    except Exception: pass
    _MN=['Enero','Febrero','Marzo','Abril','Mayo','Junio','Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']
    lab2={}
    for i,lab in enumerate(real_labels):
        pr=_parse_label(lab)
        if pr: lab2["%04d-%02d"%pr]=(i,"%s %d"%(_MN[pr[1]-1], pr[0]))
    keys=sorted(k for k in cache.keys() if k in lab2)
    months=[{"key":k,"label":lab2[k][1]} for k in keys]
    cods=set()
    for k in keys: cods|=set(cache[k].keys())
    cods|=set(real_map.keys()); cods|=set(realusd.keys())
    vpm={}; realm={}; mixm={}; reald={}
    for cod in cods:
        vr=[]; rr=[]; mm=[]; rd=[]; any_=False
        for k in keys:
            ent=cache[k].get(cod)
            vv=ent[0] if isinstance(ent,list) else ent
            mx=ent[1] if (isinstance(ent,list) and len(ent)>1) else None
            vr.append(vv); mm.append(mx)
            idx=lab2[k][0]; rv=real_map.get(cod)
            rrv=rv[idx] if (rv and idx<len(rv)) else None; rr.append(rrv)
            mo=int(k[5:7]); rud=realusd.get(cod,{}).get(mo); rd.append(rud)
            if vv is not None or rrv is not None or rud is not None: any_=True
        if any_: vpm[cod]=vr; realm[cod]=rr; mixm[cod]=mm; reald[cod]=rd
    return {"months":months,"vp":vpm,"real":realm,"mix":mixm,"reald":reald}

def build(data, out_html, tpl_path):
    tpl=open(tpl_path,encoding="utf-8").read()
    html=tpl.replace("/*__DATA__*/", json.dumps(data,ensure_ascii=False,separators=(",",":")))
    open(out_html,"w",encoding="utf-8").write(html)
    print("OK ->",out_html, round(os.path.getsize(out_html)/1024),"KB | codigos:",len(data["rows"]),"| GA con IMPO:",len(data["impo"]))

if __name__=="__main__":
    folder=os.path.dirname(os.path.abspath(__file__))
    src = sys.argv[1] if len(sys.argv)>1 else find_latest(folder)
    out = sys.argv[2] if len(sys.argv)>2 else os.path.join(folder,"index.html")
    tpl = os.path.join(folder,"plantilla.html")
    print("Leyendo:",src)
    data=extract(src)
    cp = sys.argv[3] if len(sys.argv)>3 else find_costos(folder)
    data["costos"]=load_costos(cp) if cp else {}
    print("Costos:",cp,"->",len(data["costos"]),"codigos")
    try:
        hist=build_historico(os.path.dirname(os.path.abspath(src)), data.get("real",{}), data.get("real_labels",[]), os.path.join(folder,"historico_vp.json"), read_realusd(os.path.dirname(os.path.abspath(src))))
        data["hist"]=hist
        print("Historico: meses",len(hist["months"]),"| codigos",len(hist["vp"]))
    except Exception as e:
        print("Historico: error",e); data["hist"]={"months":[],"vp":{},"real":{}}
    data.pop("real",None); data.pop("real_labels",None)
    build(data, out, tpl)
